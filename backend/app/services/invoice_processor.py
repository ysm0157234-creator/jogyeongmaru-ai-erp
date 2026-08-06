from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9가-힣]+", "", str(value or "").lower())


_QUANTITY_KEYS = (
    "quantity", "qty", "수량", "입고수량", "신고수량", "aantal", "pieces", "piece", "pcs", "units",
)


def extract_target_quantity(values: dict[str, Any]) -> int | float | None:
    """Shipment Overview 행에서 신고에 사용할 수량을 찾는다."""
    ranked: list[tuple[int, int | float]] = []
    for key, value in (values or {}).items():
        key_norm = normalize(key)
        score = 0
        for index, word in enumerate(_QUANTITY_KEYS):
            if normalize(word) == key_norm:
                score = max(score, 100 - index)
            elif normalize(word) in key_norm:
                score = max(score, 70 - index)
        if not score:
            continue
        if isinstance(value, bool):
            continue
        number: int | float | None = None
        if isinstance(value, (int, float)):
            number = value
        else:
            match = re.search(r"-?\d+(?:[,.]\d+)?", str(value or "").replace(",", ""))
            if match:
                number = float(match.group())
        if number is not None and number > 0:
            if float(number).is_integer():
                number = int(number)
            ranked.append((score, number))
    return sorted(ranked, key=lambda item: -item[0])[0][1] if ranked else None


def _matching_rows(ws, variety_name: str) -> list[int]:
    target = normalize(variety_name)
    words = [normalize(v) for v in re.split(r"\s+", variety_name) if len(normalize(v)) >= 3]
    cultivar = words[-1] if words else ""
    result: list[int] = []
    for row in range(1, ws.max_row + 1):
        joined = normalize(" ".join(str(ws.cell(row, col).value or "") for col in range(1, ws.max_column + 1)))
        if target and target in joined:
            result.append(row)
        elif len(words) >= 2 and all(word in joined for word in words[:2]):
            result.append(row)
        elif cultivar and len(cultivar) >= 5 and cultivar in joined:
            result.append(row)
    return result


def _quantity_column(ws, matched_row: int) -> int | None:
    # 품종 행 위쪽 헤더에서 수량 열을 찾는다.
    for header_row in range(max(1, matched_row - 15), matched_row):
        for col in range(1, ws.max_column + 1):
            value = normalize(ws.cell(header_row, col).value)
            if any(normalize(key) == value or normalize(key) in value for key in _QUANTITY_KEYS):
                return col

    # 헤더가 불명확하면 품종명 오른쪽의 첫 정수 셀을 수량으로 본다.
    variety_cols = []
    for col in range(1, ws.max_column + 1):
        value = normalize(ws.cell(matched_row, col).value)
        if value and any(part in value for part in normalize(str(ws.cell(matched_row, col).value)).split()):
            variety_cols.append(col)
    start = min(variety_cols) + 1 if variety_cols else 1
    for col in range(start, ws.max_column + 1):
        value = ws.cell(matched_row, col).value
        if isinstance(value, int) and 0 < value < 10_000_000:
            return col
    return None


def update_invoice_xlsx_quantity(
    source_data: bytes,
    variety_name: str,
    target_quantity: int | float | None,
    *,
    keep_vba: bool = False,
) -> bytes:
    """원본 엑셀 레이아웃을 유지하고 일치 품종 행의 수량 셀만 변경한다."""
    workbook = load_workbook(io.BytesIO(source_data), keep_vba=keep_vba)
    changed = 0
    if target_quantity is not None:
        for ws in workbook.worksheets:
            for row in _matching_rows(ws, variety_name):
                col = _quantity_column(ws, row)
                if col:
                    ws.cell(row, col).value = target_quantity
                    changed += 1
        # Excel/한셀에서 수식이 자동 재계산되도록 설정
        try:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcMode = "auto"
        except Exception:
            pass
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _number_words_same_line(page, rect):
    words = page.get_text("words")
    center_y = (rect.y0 + rect.y1) / 2
    candidates = []
    for word in words:
        x0, y0, x1, y1, text = word[:5]
        word_center = (y0 + y1) / 2
        if abs(word_center - center_y) > max(8, rect.height * 0.8):
            continue
        if x0 <= rect.x1 - 2:
            continue
        clean = str(text).replace(",", "").strip()
        if not re.fullmatch(r"\d+(?:\.\d+)?", clean):
            continue
        number = float(clean)
        if number <= 0 or number >= 10_000_000:
            continue
        # 품종명 오른쪽에서 가까운 숫자를 우선한다. 소수는 단가일 가능성이 높아 감점.
        score = (x0 - rect.x1) + (80 if "." in clean else 0)
        candidates.append((score, word, clean))
    return sorted(candidates, key=lambda item: item[0])


def update_invoice_pdf_quantity(
    source_data: bytes,
    variety_name: str,
    target_quantity: int | float | None,
) -> tuple[bytes, bool]:
    """원본 PDF 페이지를 유지하고 품종 행의 수량 숫자만 덮어쓴다."""
    if target_quantity is None:
        return source_data, False
    try:
        import fitz  # PyMuPDF
    except Exception:
        return source_data, False

    doc = fitz.open(stream=source_data, filetype="pdf")
    changed = False
    search_terms = [variety_name]
    parts = [part for part in re.split(r"\s+", variety_name.strip()) if len(part) >= 4]
    if parts:
        search_terms.append(parts[-1])

    for page in doc:
        variety_rects = []
        for term in search_terms:
            variety_rects = page.search_for(term)
            if variety_rects:
                break
        for variety_rect in variety_rects:
            candidates = _number_words_same_line(page, variety_rect)
            if not candidates:
                continue
            _, word, old_text = candidates[0]
            x0, y0, x1, y1 = word[:4]
            number_rect = fitz.Rect(x0 - 1, y0 - 1, x1 + 2, y1 + 1)
            page.add_redact_annot(number_rect, fill=(1, 1, 1))
            page.apply_redactions()
            new_text = str(int(target_quantity) if float(target_quantity).is_integer() else target_quantity)
            fontsize = max(6, min(12, (y1 - y0) * 0.85))
            page.insert_textbox(
                number_rect,
                new_text,
                fontsize=fontsize,
                fontname="helv",
                color=(0, 0, 0),
                align=0,
            )
            changed = True
            break
        if changed:
            break

    if not changed:
        doc.close()
        return source_data, False
    output = doc.tobytes(garbage=4, deflate=True)
    doc.close()
    return output, True


# 이전 코드와의 호환용 이름. 더 이상 행 삭제나 발췌본 생성은 하지 않는다.
def filter_invoice_xlsx(source_data: bytes, variety_name: str, shipment_number: str) -> bytes:
    return update_invoice_xlsx_quantity(source_data, variety_name, None)


def extract_invoice_pdf_pages(source_data: bytes, variety_name: str) -> tuple[bytes, int]:
    reader = PdfReader(io.BytesIO(source_data))
    return source_data, len(reader.pages)


def create_invoice_extract_xlsx(
    variety_name: str,
    shipment_number: str,
    shipment_values: dict[str, Any],
    source_invoice_name: str | None,
) -> bytes:
    raise LookupError("원본 Invoice가 없어 발췌본을 만들지 않았습니다. 원본 Invoice를 직접 업로드하세요.")


def workbook_contains(data: bytes, terms: list[str]) -> bool:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    targets = [normalize(term) for term in terms if term]
    return any(
        any(target and target in normalize(" ".join(str(v or "") for v in row)) for target in targets)
        for ws in wb.worksheets for row in ws.iter_rows(values_only=True)
    )


def pdf_contains(data: bytes, terms: list[str]) -> bool:
    reader = PdfReader(io.BytesIO(data))
    targets = [normalize(t) for t in terms if t]
    return any(any(t and t in normalize(page.extract_text() or "") for t in targets) for page in reader.pages)
