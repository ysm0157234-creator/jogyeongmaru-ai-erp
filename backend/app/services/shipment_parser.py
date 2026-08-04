from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook


@dataclass
class ShipmentMatch:
    sheet_name: str
    row_number: int
    description: str
    shipment: str
    values: dict[str, Any]
    source: str = "shipment_overview"


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9가-힣]+", "", str(value or "").lower())


def tokens_for_variety(variety_name: str) -> list[str]:
    raw = str(variety_name or "").strip()
    cultivar = re.sub(
        r"(?i)^\s*tulipa\s+(?:spp\.?\s*)?",
        "",
        raw,
    ).strip()

    candidates = {
        raw,
        raw.replace("Sunlover", "Sun Lover"),
        raw.replace("Sun Lover", "Sunlover"),
        raw.replace("spp.", ""),
        raw.replace("spp", ""),
        cultivar,
        cultivar.replace(" ", ""),
        "Tulipa " + cultivar if cultivar else "",
    }

    return sorted(
        {
            normalize(value)
            for value in candidates
            if len(normalize(value)) >= 4
        },
        key=len,
        reverse=True,
    )


def row_score(row_text: str, terms: list[str]) -> int:
    normalized_row = normalize(row_text)
    if not normalized_row:
        return 0

    best = 0
    for index, term in enumerate(terms):
        if term and term in normalized_row:
            best = max(best, 200 - index)

    # Sun Lover처럼 띄어쓰기 차이가 있는 경우
    compact = normalized_row.replace(" ", "")
    for index, term in enumerate(terms):
        if term and term.replace(" ", "") in compact:
            best = max(best, 180 - index)

    return best


def find_variety_in_workbook(
    data: bytes,
    variety_name: str,
) -> ShipmentMatch:
    try:
        workbook = load_workbook(
            io.BytesIO(data),
            data_only=False,
            read_only=False,
        )
    except Exception as exc:
        raise LookupError(
            f"Shipment Overview 엑셀을 열지 못했습니다: {exc}"
        ) from exc

    terms = tokens_for_variety(variety_name)
    matches: list[tuple[int, ShipmentMatch]] = []

    for sheet in workbook.worksheets:
        if sheet.max_row < 1 or sheet.max_column < 8:
            continue

        # 헤더를 찾되, 못 찾으면 1행부터 전부 검사
        header_row = 1
        for row in range(1, min(sheet.max_row, 40) + 1):
            row_headers = [
                normalize(sheet.cell(row, col).value)
                for col in range(1, min(sheet.max_column, 30) + 1)
            ]
            if any(
                any(keyword in header for keyword in (
                    "shipment",
                    "description",
                    "품종",
                    "품명",
                    "item",
                    "product",
                ))
                for header in row_headers
            ):
                header_row = row
                break

        headers = {
            col: str(sheet.cell(header_row, col).value or f"COL_{col}")
            for col in range(1, sheet.max_column + 1)
        }

        # H열은 고정 Shipment 열
        shipment_col = 8

        # 헤더 바로 다음 행부터 검색하되, 헤더 탐지가 잘못돼도 전체 행을 놓치지 않음
        start_row = max(1, header_row + 1)

        for row in range(start_row, sheet.max_row + 1):
            cell_values = [
                sheet.cell(row, col).value
                for col in range(1, sheet.max_column + 1)
            ]
            row_text = " | ".join(
                str(value or "")
                for value in cell_values
            )

            score = row_score(row_text, terms)
            if not score:
                continue

            shipment = str(
                sheet.cell(row, shipment_col).value or ""
            ).strip()

            # H열이 비어 있으면 후보로는 잡되 우선순위를 낮춤
            if shipment:
                score += 100

            description = next(
                (
                    str(value).strip()
                    for value in cell_values
                    if value
                    and any(
                        term in normalize(value)
                        for term in terms
                    )
                ),
                row_text,
            )

            values = {
                headers[col]: sheet.cell(row, col).value
                for col in range(1, sheet.max_column + 1)
                if sheet.cell(row, col).value not in (None, "")
            }
            values["Shipment(H열)"] = shipment

            matches.append(
                (
                    score,
                    ShipmentMatch(
                        sheet_name=sheet.title,
                        row_number=row,
                        description=description,
                        shipment=shipment,
                        values=values,
                    ),
                )
            )

    if not matches:
        raise LookupError(
            f"Shipment Overview의 모든 시트와 행에서 "
            f"'{variety_name}' 품종을 찾지 못했습니다."
        )

    matches.sort(
        key=lambda item: (
            item[0],
            bool(item[1].shipment),
        ),
        reverse=True,
    )
    best = matches[0][1]

    if not best.shipment:
        raise LookupError(
            f"품종은 찾았지만 같은 행 H열 Shipment 값이 비어 있습니다. "
            f"(시트: {best.sheet_name}, 행: {best.row_number})"
        )

    return best
